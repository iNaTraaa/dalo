import os
import time
import random
import openpyxl
import pyautogui
import pyperclip
import sys 
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# cấu hình màu
VANG_FILL = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
DO_FILL = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

MESSAGE_TEMPLATE = """Chào em, Thầy đến từ Khoa Công nghệ Thông tin – Trường Đại học Nguyễn Tất Thành đây 🤝

Thầy cô nhắn để hỗ trợ em tư vấn chọn ngành học phù hợp hơn với bản thân. Em chỉ cần trả lời bằng một số trong các lựa chọn sau nhé:

1. Rất quan tâm
2. Quan tâm
3. Đang cân nhắc
4. Muốn nhận thêm thông tin
5. Đã đăng ký / đang làm hồ sơ
6. Phụ huynh muốn cùng trao đổi
7. Chưa quan tâm / đã chọn trường khác
8. Khác: (nội dung em muốn chia sẻ)

Thầy cô Khoa CNTT luôn sẵn sàng hỗ trợ em!"""

def setup_profile():  #Tự động tạo và trả về đường dẫn thư mục Profile ngay tại nơi đặt code

    current_dir = os.path.dirname(os.path.abspath(__file__))
    profile_path = os.path.join(current_dir, "ZaloToolData")
    
    if not os.path.exists(profile_path):
        os.makedirs(profile_path)
        print(f"Đã khởi tạo thư mục dữ liệu tại: {profile_path}")
    
    return profile_path

def update_excel(sheet, row, col6, col7, color, path, wb):  #Cập nhật dữ liệu vào Excel
    sheet.cell(row=row, column=6).value = col6
    sheet.cell(row=row, column=7).value = col7
    for col in range(1, 9):
        sheet.cell(row=row, column=col).fill = color
    wb.save(path)

def check_blocked_status(driver):   #Kiểm tra các câu thông báo chặn của Zalo
    keywords = ["không nhận tin nhắn", "người lạ", "không thể nhận"]
    for word in keywords:
        if driver.find_elements(By.XPATH, f"//*[contains(text(), '{word}')]"):
            return True
    return False

def process_send(driver, phone):  #Xử lý logic gửi tin và trả về kết quả 
    try:
        # click vào body để focus
        driver.find_element(By.TAG_NAME, 'body').click()
        time.sleep(1)

        search_box = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "contact-search-input"))
        )
        search_box.click()
        search_box.clear()
        search_box.send_keys(phone)
        time.sleep(2)
        search_box.send_keys(Keys.ENTER)
        time.sleep(3)

        # Không tìm thấy tài khoản
        not_found_msgs = ["chưa đăng ký", "Không tìm thấy", "không cho phép"]
        for msg in not_found_msgs:
            if driver.find_elements(By.XPATH, f"//*[contains(text(), '{msg}')]"):
                pyautogui.press('esc')
                return 3

        # Click nút Nhắn tin nếu có
        btns = driver.find_elements(By.XPATH, "//div[contains(text(), 'Nhắn tin')] | //span[contains(text(), 'Nhắn tin')]")
        if btns: btns[0].click()
        time.sleep(2)

        # Kiểm tra chặn trước khi gửi
        if check_blocked_status(driver):
            pyautogui.press('esc')
            return 2

        # Thực hiện gửi tin nhắn
        pyperclip.copy(MESSAGE_TEMPLATE)
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(1)
        pyautogui.press('enter')
        time.sleep(3)

        # Kiểm tra lại lần nữa sau khi gửi vì Zalo có thể hiện thông báo sau khi nhấn Enter
        if check_blocked_status(driver):
            return 2
        # Kiểm tra coi có ban không
        if driver.find_elements(By.XPATH, "//*[text()='Không thể nhận tin nhắn từ bạn.']"):
            return 4

        return 1
    except Exception as e:
        print(f"Lỗi: {e}")
        return 0

def main():
    print("Bắt đầu chạy tool")
    excel_path = input("Nhập đường dẫn file excel: ").strip().replace('"', '')
    
    try:
        start_stt = int(input("Nhập STT bắt đầu (Cột A): "))
        end_stt = int(input("Nhập STT kết thúc (Cột A): "))
    except:
        print("Lỗi: STT phải là số!")
        return

    # Tự động lấy Profile
    profile_path = setup_profile()

    # Mở Excel
    try:
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
    except Exception as e:
        print(f"Không mở được file Excel: {e}")
        return

    # Khởi tạo trình duyệt
    options = Options()
    options.add_argument(f"user-data-dir={profile_path}")
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    
    driver.get("https://chat.zalo.me")
    print("Nếu chưa đăng nhập, hãy quét mã QR ngay...")
    time.sleep(25) # Đợi đăng nhập/tải trang

    processed = 0
    for row in range(2, sheet.max_row + 1):
        stt_val = sheet.cell(row=row, column=1).value
        try:
            current_stt = int(stt_val)
            if not (start_stt <= current_stt <= end_stt): continue
        except: continue

        phone = str(sheet.cell(row=row, column=4).value).strip()
        if not phone or phone == "None": continue

        print(f"[Đang xử lý STT {current_stt}: {phone}")
        
        result = process_send(driver, phone)


        if result == 4:
            print("CẢNH BÁO: TÀI KHOẢN ĐANG BỊ ZALO BAN ")
            # Cập nhật trạng thái lỗi vào Excel cho dòng hiện tại trước khi nghỉ
            update_excel(sheet, row, "BỊ CHẶN", "Tài khoản bị Zalo Ban tính năng", DO_FILL, excel_path, wb)
            driver.quit()
            sys.exit() # Dừng toàn bộ chương trình
        elif result == 1:
            print(" ->Thành công")
            update_excel(sheet, row, "Đã gửi", "Đợi trả lời", VANG_FILL, excel_path, wb)
        elif result == 2:
            print(" ->Bị chặn người lạ")
            update_excel(sheet, row, "Đã gửi", "Không nhận tin nhắn người lạ", DO_FILL, excel_path, wb)
        elif result == 3:
            print(" ->Không tìm thấy SĐT")
            update_excel(sheet, row, "Không thể gửi", "Chặn người lạ tìm số", DO_FILL, excel_path, wb)
         
        processed += 1
        
        # Nghỉ tránh spam 
        if processed % 15 == 0:
            print("Nghỉ 3 phút ")
            time.sleep(180)
        else:
            delay = random.randint(20, 30)
            print(f"Đợi {delay}s...")
            time.sleep(delay)

    print("Hoàn tất")
    driver.quit()

if __name__ == "__main__":
    main()